from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from transformers import AutoTokenizer, AutoModelForSequenceClassification, pipeline
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import io
from datetime import datetime


# List to store collected chat data
collected_data = []

# Load pre-trained model and tokenizer for sentiment analysis
model_name = "cardiffnlp/twitter-roberta-base-sentiment"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForSequenceClassification.from_pretrained(model_name)

# Create sentiment analysis pipeline
sentiment_pipeline = pipeline("sentiment-analysis", model=model,tokenizer=tokenizer)



# Mapping of model labels to human-readable sentiments
label_map = {
    'LABEL_0': 'Negative',
    'LABEL_1': 'Neutral',
    'LABEL_2': 'Positive'
}

# Initialize Selenium WebDriver (make sure the Chrome WebDriver is installed and in PATH)
driver = webdriver.Chrome()

# Open the YouTube video with live chat
driver.get("https://www.youtube.com/watch?v=QFbD8r7JYZo")

# Wait for the chat frame to load and switch to it
WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "chatframe")))

# Record start time for data collection duration
start_time = time.time()

try:
    # Loop to continuously collect chat messages
    while True:
        # Find all live chat messages in the YouTube live stream
        messages = driver.find_elements(By.CSS_SELECTOR, "yt-live-chat-text-message-renderer")
        
        # Iterate over the collected messages
        for message in messages:
            content = message.find_element(By.CSS_SELECTOR, "#message").text
            current_time = datetime.now().strftime('%H:%M:%S:%f')[:-3]
            # Perform sentiment analysis on the message content
            sentiment_label = sentiment_pipeline(content)[0]['label']
            # Convert the label into a human-readable sentiment
            sentiment = label_map.get(sentiment_label, "Unknown")

            # Append the message content, timestamp, and sentiment to the collected data list
            collected_data.append({
                "Time": current_time,
                "Content": content,
                "Sentiment": sentiment
            })

            # Print the message and sentiment for real-time feedback
            print(f"Time: {current_time}, Content: {content}, Sentiment: {sentiment}")

        # Wait for 5 seconds before collecting new messages to avoid overloading
        time.sleep(5)

        # Stop collecting messages after 30 seconds
        elapsed_time = time.time() - start_time
        if elapsed_time > 30:
            print("30 seconds have passed. Stopping data extraction.")
            break
finally:
    driver.switch_to.default_content()  # Switch back to the main content frame
    driver.quit()  # Close the WebDriver

# Convert collected data into a Pandas DataFrame
df = pd.DataFrame(collected_data)

# Print the total number of responses collected
total_responses = len(df)
print(f"Total responses collected: {total_responses}")

# Define colors for each sentiment for visualizations and Excel formatting
sentiment_colors = {
    'Positive': '#66c2a5',  # Green for positive
    'Neutral': '#fc8d62',   # Orange for neutral
    'Negative': '#8da0cb'   # Blue for negative
}

# Count the occurrences of each sentiment
sentiment_counts = df['Sentiment'].value_counts()
labels = sentiment_counts.index
sizes = sentiment_counts.values
colors = [sentiment_colors[label] for label in labels]

# Generate a pie chart showing sentiment distribution
plt.figure(figsize=(6,6))
plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140, colors=colors)
plt.axis('equal')  # Ensure the pie chart is a perfect circle
plt.title('Sentiment Analysis of YouTube Live Chat Messages')

# Save the pie chart to a buffer (in memory) to add it later in the Excel file
pie_chart_buffer = io.BytesIO()
plt.savefig(pie_chart_buffer, format='png')
plt.close()

# Define the path to save the Excel report
file_path = r'G:\My Drive\YoutubeLiveMessageSentiments.xlsx'

# Create an Excel writer using OpenPyXL to write data to the specified Excel file
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    # Write the DataFrame (chat data) to the "Chat Data" sheet
    df.to_excel(writer, sheet_name='Chat Data', index=False)

    # Get the workbook and worksheet objects for further modifications
    workbook = writer.book
    worksheet = writer.sheets['Chat Data']

    # Apply color formatting to the sentiment column based on sentiment value
    for row in range(2, len(df) + 2):  # Start from row 2 (row 1 is headers)
        sentiment_value = worksheet.cell(row=row, column=3).value  # Column 3 is "Sentiment"
        if sentiment_value in sentiment_colors:
            color_hex = sentiment_colors[sentiment_value].replace('#', '')
          
  # Fill the cell with the corresponding color
            worksheet.cell(row=row, column=3).fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    # Adjust the width of the columns for better readability
    worksheet.column_dimensions[get_column_letter(1)].width = 13  # "Time" column
    worksheet.column_dimensions[get_column_letter(2)].width = 37  # "Content" column
    worksheet.column_dimensions[get_column_letter(3)].width = 10  # "Sentiment" column

    # Enable text wrapping for cells in the DataFrame range
    for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True)

    # Create a new worksheet to add the sentiment pie chart
    chart_worksheet = workbook.create_sheet('Sentiment Pie Chart')

    # Insert the pie chart image into the new worksheet
    pie_chart_image = Image(pie_chart_buffer)
    chart_worksheet.add_image(pie_chart_image, 'A1')

# Print a message to indicate the Excel report has been generated successfully
print(f"Excel report generated: {file_path}")
